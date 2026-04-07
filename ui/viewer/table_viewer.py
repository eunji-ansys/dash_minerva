import os
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import dash_bootstrap_components as dbc
import pandas as pd
from dash import html
from openpyxl import load_workbook


def sanitize_xlsx_for_preview(file_path: str | Path) -> str:
    temp_dir = tempfile.mkdtemp(prefix="xlsx_preview_")
    sanitized_path = os.path.join(temp_dir, Path(file_path).name)

    with zipfile.ZipFile(file_path, "r") as zin:
        with zipfile.ZipFile(sanitized_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)

                if item.filename == "xl/workbook.xml":
                    try:
                        root = ET.fromstring(data)
                        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                        defined_names = root.find("main:definedNames", ns)
                        if defined_names is not None:
                            root.remove(defined_names)
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                    except Exception:
                        pass

                zout.writestr(item, data)

    return sanitized_path


def render_table_viewer(file_path: str | Path, max_rows: int = 200):
    ext = Path(file_path).suffix.lower()

    try:
        if ext == ".csv":
            df = pd.read_csv(file_path, nrows=max_rows)
        elif ext == ".xlsx":
            preview_path = sanitize_xlsx_for_preview(file_path)
            wb = load_workbook(preview_path, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]

            rows = list(ws.iter_rows(values_only=True, max_row=max_rows + 1))
            if not rows:
                return html.Div("Empty worksheet.", className="text-muted")

            header = rows[0]
            data_rows = rows[1:]
            normalized_header = [
                str(col).strip() if col not in (None, "") else f"Column {i+1}"
                for i, col in enumerate(header)
            ]
            df = pd.DataFrame(data_rows, columns=normalized_header)
        else:
            return html.Div("Unsupported table format.", className="text-muted")

        return html.Div(
            [
                html.Div(
                    f"Previewing first {min(len(df), max_rows)} row(s)",
                    className="text-muted small mb-2",
                ),
                dbc.Table.from_dataframe(
                    df,
                    striped=True,
                    bordered=True,
                    hover=True,
                    size="sm",
                    responsive=True,
                    className="mb-0",
                ),
            ],
            style={"maxHeight": "70vh", "overflow": "auto"},
        )
    except Exception:
        return html.Div(
            [
                html.Div("Preview not available for this Excel file.", className="fw-semibold"),
                html.Div("This file contains unsupported Excel metadata.", className="text-muted small"),
                html.Div("Please download and open in Excel.", className="text-muted small mt-2"),
            ],
            className="p-3",
        )